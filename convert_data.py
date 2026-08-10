"""
将 C_memorize 目录下所有 xlsx 文件转换为统一的 words.json。
支持一个文件包含多个词条（以标题行分割）。

Excel 结构（每个词）：
  标题行: "序号．汉字（拼音）" 或 "序号．汉字 (拼音）"
  空行:   分隔
  表头行: 词性 | 词义 | 例句 | 篇名
  数据行: 词性列有纵向合并单元格

转换规则：
  - 词性合并单元格 → 向下填充
  - 词义为 None 但例句有值 → 沿用上一个非空词义
  - 输出 JSON: [{word, pinyin, index, senses: [{meaning, example, source}]}]
"""

import json
import re
import os
from glob import glob

import openpyxl


TITLE_RE = re.compile(r'(\d+)[．.]\s*(.+?)[（(](.+?)[）)]')


def parse_title(text):
    """从 '1．哀（āi）' 解析出序号、字、拼音"""
    m = TITLE_RE.match(str(text).strip())
    if m:
        return int(m.group(1)), m.group(2).strip(), m.group(3).strip()
    return None, None, None


def is_title_row(cells):
    """检测是否为词条标题行"""
    if cells and cells[0]:
        return TITLE_RE.match(str(cells[0]).strip()) is not None
    return False


def is_header_row(cells):
    """检测是否为表头行"""
    if cells and len(cells) >= 4:
        return str(cells[0]).strip() == '词性'
    return False


def parse_word_block(title_cells, data_rows):
    """解析一个词条的数据块，返回 dict 或 None"""
    title = str(title_cells[0]) if title_cells[0] else ""
    index, word, pinyin = parse_title(title)
    if not word:
        return None

    senses = []
    current_ciyi = ""

    for row in data_rows:
        ciyi = str(row[1]).strip() if row[1] else ""
        liju = str(row[2]).strip() if row[2] else ""
        pianming = str(row[3]).strip() if row[3] else ""

        if ciyi:
            current_ciyi = ciyi

        if liju and current_ciyi:
            liju_highlighted = liju.replace(word, f"<b>{word}</b>")
            senses.append({
                "meaning": current_ciyi,
                "example": liju_highlighted,
                "source": pianming
            })

    return {
        "word": word,
        "pinyin": pinyin,
        "index": index,
        "senses": senses
    }


def convert_xlsx(filepath):
    """转换一个 xlsx 文件，可能包含多个词条"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # 读取所有行
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = list(row)
        all_rows.append(cells)

    words = []
    i = 0
    while i < len(all_rows):
        row = all_rows[i]

        if is_title_row(row):
            # 找到标题行，收集后续数据行直到下一个标题行
            title_cells = row
            i += 1
            data_rows = []

            while i < len(all_rows):
                next_row = all_rows[i]
                if is_title_row(next_row):
                    break
                if is_header_row(next_row):
                    # 跳过表头行
                    i += 1
                    continue
                # 收集非全空的数据行
                if any(c is not None for c in next_row):
                    data_rows.append(next_row)
                i += 1

            entry = parse_word_block(title_cells, data_rows)
            if entry:
                words.append(entry)
                print(f"  → {entry['word']}（{entry['pinyin']}），{len(entry['senses'])}个义项")
        else:
            i += 1

    return words


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    files = sorted(glob(os.path.join(base_dir, "*.xlsx")))

    if not files:
        print("未找到 xlsx 文件！")
        return

    all_words = []
    for f in files:
        fname = os.path.basename(f)
        # 跳过临时文件
        if fname.startswith('~$'):
            continue
        print(f"处理: {fname}")
        words = convert_xlsx(f)
        all_words.extend(words)

    # 按 index 排序
    all_words.sort(key=lambda w: w["index"])

    output_path = os.path.join(base_dir, "words.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_words, f, ensure_ascii=False, indent=2)

    print(f"\n共转换 {len(all_words)} 个词，输出至 {output_path}")


if __name__ == "__main__":
    main()
